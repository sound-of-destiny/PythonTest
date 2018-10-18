from openpyxl import load_workbook
from influxdb import InfluxDBClient
import os

client = InfluxDBClient('localhost', 8086, 'admin', 'admin', 'waterqualify')

files = os.listdir('/home/schong/Desktop/data')

for XLSX_FIlE in files:
    wb = load_workbook('/home/schong/Desktop/data/' + XLSX_FIlE)
    sheets = wb.sheetnames
    for sheet in sheets:
        booksheet = wb[sheet]
        rows = booksheet.rows
        s = 0
        for row in rows:
            s += 1
            if s < 4 :
                continue
        
            time = booksheet['A' + str(s)].value
            tempMin = booksheet['B' + str(s)].value
            tempMax = booksheet['C' + str(s)].value
            tempAve = booksheet['D' + str(s)].value

            PHMin = booksheet['E' + str(s)].value
            PHMax = booksheet['F' + str(s)].value
            PHAve = booksheet['G' + str(s)].value

            turbMin = booksheet['H' + str(s)].value
            turbMax = booksheet['I' + str(s)].value
            turbAve = booksheet['J' + str(s)].value

            presMin = booksheet['K' + str(s)].value
            presMax = booksheet['L' + str(s)].value
            presAve = booksheet['M' + str(s)].value

            chloMin = booksheet['N' + str(s)].value
            chloMax = booksheet['O' + str(s)].value
            chloAve = booksheet['P' + str(s)].value

            condMin = booksheet['Q' + str(s)].value
            condMax = booksheet['R' + str(s)].value
            condAve = booksheet['S' + str(s)].value

            orgcMin = booksheet['T' + str(s)].value
            orgcMax = booksheet['U' + str(s)].value
            orgcAve = booksheet['V' + str(s)].value

            json_body = [{
                "measurement": sheet,
                "tags": {
                },
                "time": time,
                "fields": {
                        "TempMin": tempMin,
                        "TempMax": tempMax,
                        "TempAve": tempAve,
                        "PHMin": PHMin,
                        "PHMax": PHMax,
                        "PHAve": PHAve,
                        "TurbMin": turbMin,
                        "TurbMax": turbMax,
                        "TurbAve": turbAve,
                        "PresMin": presMin,
                        "PresMax": presMax,
                        "PresAve": presAve,
                        "ChloMin": chloMin,
                        "ChloMax": chloMax,
                        "ChloAve": chloAve,
                        "CondMin": condMin,
                        "CondMax": condMax,
                        "CondAve": condAve,
                        "OrgcMin": orgcMin,
                        "OrgcMax": orgcMax,
                        "OrgcAve": orgcAve
                }
                }]

            client.write_points(json_body)


        
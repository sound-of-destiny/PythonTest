from openpyxl import load_workbook
from influxdb import InfluxDBClient
from sklearn.cluster import DBSCAN
from sklearn.preprocessing import StandardScaler
import numpy as np
import os

client = InfluxDBClient('localhost', 8086, 'admin', 'admin', 'waterqualify')

files = os.listdir('/home/schong/Desktop/data')

#X = np.array([0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0])
#X = np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0])
X = np.array([0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0,0.0])

time = []
sheetname = []

for XLSX_FIlE in files:
    wb = load_workbook('/home/schong/Desktop/data/' + XLSX_FIlE)
    sheets = wb.sheetnames
    for sheet in sheets:
        booksheet = wb[sheet]
        rows = booksheet.rows
        s = 0
        for row in rows:
            s += 1
            if s < 4 :
                continue
            time.append(booksheet['A' + str(s)].value)
            sheetname.append(sheet)

            tempMin = booksheet['B' + str(s)].value
            tempMax = booksheet['C' + str(s)].value
            tempAve = booksheet['D' + str(s)].value

            PHMin = booksheet['E' + str(s)].value
            PHMax = booksheet['F' + str(s)].value
            PHAve = booksheet['G' + str(s)].value

            turbMin = booksheet['H' + str(s)].value
            turbMax = booksheet['I' + str(s)].value
            turbAve = booksheet['J' + str(s)].value

            presMin = booksheet['K' + str(s)].value
            presMax = booksheet['L' + str(s)].value
            presAve = booksheet['M' + str(s)].value

            chloMin = booksheet['N' + str(s)].value
            chloMax = booksheet['O' + str(s)].value
            chloAve = booksheet['P' + str(s)].value

            condMin = booksheet['Q' + str(s)].value
            condMax = booksheet['R' + str(s)].value
            condAve = booksheet['S' + str(s)].value

            orgcMin = booksheet['T' + str(s)].value
            orgcMax = booksheet['U' + str(s)].value
            orgcAve = booksheet['V' + str(s)].value

            #tmp = np.array([tempMin,tempMax,PHMin,PHMax,turbMin,turbMax,presMin,presMax,chloMin,chloMax,condMin,condMax,orgcMin,orgcMax])
            tmp = np.array([tempMin,tempMax,PHMin,PHMax,turbMin,turbMax,presMin,presMax,chloMin,chloMax,condMin,condMax])
            X = np.vstack((X,tmp))

X = X[1:]
X = StandardScaler().fit_transform(X)
clustering = DBSCAN(eps=0.3, min_samples=5).fit(X)
i = 0
for label in clustering.labels_:
    print(label,time[i],sheetname[i])
    i = i + 1

            
